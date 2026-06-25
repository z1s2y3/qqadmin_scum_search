"""
SCUM物品代码查询插件 - AstrBot版本
支持模糊搜索物品名、中文名称、分类
数据来源: scum.xlsx
"""

import os
from typing import List, Dict

import openpyxl
from astrbot.api.event import filter, AstrMessageEvent
from astrbot.api.star import Star, register
from astrbot.api import logger

SCUM_ITEMS: List[Dict[str, str]] = []
MAIN_CATEGORIES: List[str] = []
SUB_CATEGORIES: Dict[str, List[str]] = {}
MAIN_CATEGORY_COUNTS: Dict[str, int] = {}
SUB_CATEGORY_COUNTS: Dict[str, Dict[str, int]] = {}

DEFAULT_EXCEL = "scum.xlsx"
DEFAULT_MAX_RESULTS = 10
SEPARATOR = "──────────────────────────────────────────"
HEADER_LINE = "══════════════════════════════════════════"

CATEGORY_ICONS = {
    "载具": "🚗",
    "丧尸": "🧟",
}

SUBCATEGORY_ICONS = {
    "武器": "🔫",
    "近战": "🔫",
    "弹药": "💣",
    "食物": "🍖",
    "医疗": "💊",
}


def _get_category_icon(main_cat: str, sub_cat: str) -> str:
    if main_cat in CATEGORY_ICONS:
        return CATEGORY_ICONS[main_cat]
    for key, icon in SUBCATEGORY_ICONS.items():
        if key in sub_cat:
            return icon
    if "工具" in main_cat:
        return "🔧"
    return "📦"


def _safe_str(value) -> str:
    return str(value).strip() if value else ""


def load_items(excel_path: str = None) -> None:
    if excel_path is None:
        excel_path = os.path.join(os.path.dirname(__file__), DEFAULT_EXCEL)
    else:
        excel_path = os.path.join(os.path.dirname(__file__), excel_path)

    logger.info(f"[SCUM插件] 尝试加载Excel: {excel_path}")

    if not os.path.exists(excel_path):
        logger.error(f"[SCUM插件] Excel文件不存在: {excel_path}")
        return

    SCUM_ITEMS.clear()
    MAIN_CATEGORIES.clear()
    SUB_CATEGORIES.clear()
    MAIN_CATEGORY_COUNTS.clear()
    SUB_CATEGORY_COUNTS.clear()

    try:
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        logger.info(f"[SCUM插件] Excel文件已打开，共 {sheet.max_row} 行")

        headers = [sheet.cell(row=1, column=i).value for i in range(1, sheet.max_column + 1)]
        logger.info(f"[SCUM插件] 表头: {headers}")

        for row in range(2, sheet.max_row + 1):
            try:
                code = _safe_str(sheet.cell(row=row, column=1).value)
                name = _safe_str(sheet.cell(row=row, column=2).value)
                spawn_cmd = _safe_str(sheet.cell(row=row, column=3).value)
                main_cat = _safe_str(sheet.cell(row=row, column=4).value)
                sub_cat = _safe_str(sheet.cell(row=row, column=5).value)

                if not code:
                    continue

                main_cat = main_cat or "物品"
                sub_cat = sub_cat or "其他"

                item = {
                    "main_category": main_cat,
                    "sub_category": sub_cat,
                    "code": code,
                    "name": name,
                    "spawn_cmd": spawn_cmd,
                }
                SCUM_ITEMS.append(item)

                if main_cat not in MAIN_CATEGORIES:
                    MAIN_CATEGORIES.append(main_cat)
                    MAIN_CATEGORY_COUNTS[main_cat] = 0
                MAIN_CATEGORY_COUNTS[main_cat] += 1

                if main_cat not in SUB_CATEGORIES:
                    SUB_CATEGORIES[main_cat] = []
                    SUB_CATEGORY_COUNTS[main_cat] = {}
                if sub_cat not in SUB_CATEGORIES[main_cat]:
                    SUB_CATEGORIES[main_cat].append(sub_cat)
                    SUB_CATEGORY_COUNTS[main_cat][sub_cat] = 0
                SUB_CATEGORY_COUNTS[main_cat][sub_cat] += 1
            except Exception as e:
                logger.error(f"[SCUM插件] 读取第{row}行失败: {e}")

        logger.info(f"[SCUM插件] 已加载 {len(SCUM_ITEMS)} 个物品")
        if SCUM_ITEMS:
            logger.info(f"[SCUM插件] 示例数据: {SCUM_ITEMS[0]}")
    except Exception as e:
        logger.error(f"[SCUM插件] 加载Excel失败: {e}")
        import traceback
        logger.error(traceback.format_exc())


def _calc_match_score(item: Dict[str, str], keyword: str) -> int:
    name = item["name"].lower()
    code = item["code"].lower()
    main_cat = item["main_category"].lower()
    sub_cat = item["sub_category"].lower()

    if name == keyword or code == keyword:
        return 100
    if name.startswith(keyword) or code.startswith(keyword):
        return 50
    if keyword in name or keyword in code:
        return 20
    if sub_cat == keyword:
        return 15
    if main_cat == keyword:
        return 10
    if keyword in sub_cat:
        return 8
    if keyword in main_cat:
        return 5
    return 0


def _item_matches_all(item: Dict[str, str], keywords: List[str]) -> int:
    total_score = 0
    for kw in keywords:
        if not kw:
            continue
        score = _calc_match_score(item, kw)
        if score == 0:
            return 0
        total_score += score
    return total_score


def fuzzy_search(keyword: str) -> List[Dict[str, str]]:
    keyword = keyword.strip()
    if not keyword:
        return []

    keywords = [kw.lower() for kw in keyword.split() if kw]

    scored = []
    for item in SCUM_ITEMS:
        score = _item_matches_all(item, keywords)
        if score > 0:
            scored.append((score, item))

    scored.sort(key=lambda x: (-x[0], len(x[1]["name"]), x[1]["name"]))
    return [item for _, item in scored]


logger.info("[SCUM插件] 正在初始化...")
load_items()
logger.info(f"[SCUM插件] 初始化完成，共 {len(SCUM_ITEMS)} 个物品")


@register("SCUM物品查询", "AstrBot", "查询SCUM游戏物品代码", "1.0.0")
class ScumItemSearch(Star):
    def __init__(self, context, config=None):
        super().__init__(context)
        self.config = config or {}

    async def on_load(self):
        try:
            logger.info("[SCUM插件] 开始加载SCUM物品数据...")
            excel_path = self.config.get("excel_path", None)
            logger.info(f"[SCUM插件] 从配置读取excel_path: {excel_path}")
            load_items(excel_path)
            logger.info(f"[SCUM插件] SCUM物品加载完成，共 {len(SCUM_ITEMS)} 个物品")
        except Exception as e:
            logger.error(f"[SCUM插件] 加载SCUM物品数据失败: {e}")
            import traceback
            logger.error(traceback.format_exc())

    def _get_spawn_command(self, item: Dict[str, str]) -> str:
        if item.get("spawn_cmd"):
            return item["spawn_cmd"]
        return f"#spawnitem {item['code']} 1"

    def _get_full_magazine_cmd(self, spawn_cmd: str) -> str:
        return f"{spawn_cmd} StackCount 100"

    def _format_item(self, item: Dict[str, str]) -> str:
        name = item["name"]
        code = item["code"]
        main_cat = item["main_category"]
        sub_cat = item["sub_category"]
        spawn_cmd = self._get_spawn_command(item)
        display_name = name if name else code
        cat_icon = _get_category_icon(main_cat, sub_cat)

        result = (
            f"{cat_icon} 【{display_name}】\n"
            f"   └─ {main_cat} > {sub_cat}\n"
            f"   ├─ 物品代码: {code}\n"
            f"   ├─ 刷取指令: 🎮 {spawn_cmd}\n"
        )

        if sub_cat == "弹夹" or sub_cat == "弹匣":
            full_cmd = self._get_full_magazine_cmd(spawn_cmd)
            result += f"   ├─ 满弹夹指令: 🎮 {full_cmd}\n"

        result += f"{SEPARATOR}\n"
        return result

    def _build_pagination_info(self, keyword: str, page: int, total_pages: int, total: int, cmd: str = "物品查询") -> str:
        info = f"📊 共找到 {total} 个结果"
        if total_pages > 1:
            info += f" | 📄 第 {page}/{total_pages} 页"
            if page < total_pages:
                info += f"\n💡 输入 /{cmd} {keyword} {page + 1} 查看下一页"
        return info

    def _build_category_list(self) -> str:
        response = "📦 物品分类列表\n\n"
        for main_cat in MAIN_CATEGORIES:
            count = MAIN_CATEGORY_COUNTS.get(main_cat, 0)
            response += f"【{main_cat}】({count}个)\n"
            if main_cat in SUB_CATEGORIES:
                for sub_cat in SUB_CATEGORIES[main_cat]:
                    sub_count = SUB_CATEGORY_COUNTS.get(main_cat, {}).get(sub_cat, 0)
                    response += f"  └─ {sub_cat} ({sub_count}个)\n"
            response += "\n"
        response += "💡 使用: /物品分类 <大分类名> 查看该分类下的所有物品\n"
        response += "示例: /物品分类 工具"
        return response

    def _get_category_items(self, category: str) -> List[Dict[str, str]]:
        category_lower = category.lower().strip()
        results = []
        for item in SCUM_ITEMS:
            if item["main_category"].lower() == category_lower:
                results.append(item)
        return sorted(results, key=lambda x: (x["sub_category"], x["name"]))

    @filter.command("物品分类")
    async def category_search(self, event: AstrMessageEvent):
        args = event.message_str.strip().split()
        if len(args) < 2:
            yield event.plain_result(self._build_category_list())
            return

        category = " ".join(args[1:])
        page = 1

        if len(args) >= 3 and args[-1].isdigit():
            page = int(args[-1])
            category = " ".join(args[1:-1])

        items = self._get_category_items(category)

        if not items:
            yield event.plain_result(f"❌ 未找到「{category}」分类")
            return

        max_results = self.config.get("max_results", DEFAULT_MAX_RESULTS)
        total_pages = (len(items) + max_results - 1) // max_results

        if page > total_pages:
            yield event.plain_result(f"❌ 没有第 {page} 页，共 {total_pages} 页")
            return

        start_idx = (page - 1) * max_results
        end_idx = start_idx + max_results
        page_items = items[start_idx:end_idx]

        response = f"📦 {category} 分类物品\n"
        response += f"{HEADER_LINE}\n"
        for item in page_items:
            response += self._format_item(item)

        response += self._build_pagination_info(category, page, total_pages, len(items), "物品分类")
        yield event.plain_result(response)

    @filter.command("物品查询")
    async def item_search(self, event: AstrMessageEvent):
        args = event.message_str.strip().split()
        if len(args) < 2:
            response = (
                "❌ 请输入搜索关键词\n\n"
                "💡 使用: /物品查询 <关键词>\n"
                "示例: /物品查询 ak47\n\n"
                "支持搜索: 物品名、中文名称、分类"
            )
            yield event.plain_result(response)
            return

        keyword = " ".join(args[1:])
        page = 1

        if len(args) >= 3 and args[-1].isdigit():
            page = int(args[-1])
            keyword = " ".join(args[1:-1])

        results = fuzzy_search(keyword)

        if not results:
            yield event.plain_result(f"❌ 未找到「{keyword}」相关物品")
            return

        max_results = self.config.get("max_results", DEFAULT_MAX_RESULTS)
        total_pages = (len(results) + max_results - 1) // max_results

        if page > total_pages:
            yield event.plain_result(f"❌ 没有第 {page} 页，共 {total_pages} 页")
            return

        start_idx = (page - 1) * max_results
        end_idx = start_idx + max_results
        page_results = results[start_idx:end_idx]

        response = f"{HEADER_LINE}\n"
        for item in page_results:
            response += self._format_item(item)

        response += self._build_pagination_info(keyword, page, total_pages, len(results))
        yield event.plain_result(response)

    async def terminate(self):
        logger.info("[SCUM插件] SCUM物品查询插件已卸载")
